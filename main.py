import copy
import itertools
import random
import tkinter as tk

from openpyxl import load_workbook

import xlsx_load

# todo: 以下の二つを読み込むファイルに合わせる
FILENAME = "word_list"  # 読み込むファイルの名前
SHEETNAME = "Sheet1"  # 読み込むシートの名前
WORD = 0
DESCRIPTION = 1
TRY = 2
CORRECT = 3
RATE = 4


class App(tk.Frame):
    sheet = xlsx_load.load(FILENAME + ".xlsx", SHEETNAME)  # csvロード
    # 最初にすべての要素が含まれたリストを作成しておき、そこから表示したものを削除していく
    row_list = list(range(0, len(sheet)))
    row_choice = copy.deepcopy(row_list)
    print("Row list is " + str(row_list) + ".")
    permission_error = False

    def __init__(self, master):
        super().__init__(master)
        self.pack()
        self.create_widgets()
        master.geometry("320x180")
        master.title("暗記カード")
        self.master = master
        self.next()

    def create_widgets(self):
        self.frame_main = tk.Frame(self)
        self.frame_main.grid(row=0, column=0, sticky="nsew")

        self.frame_text = tk.Frame(self.frame_main)  # 問題文を表示するフレーム
        self.frame_text.pack(side="top")

        self.frame_entry = tk.Frame(self.frame_main)  # テキストボックスを表示するフレーム
        self.frame_entry.pack(side="top")

        self.frame_button = tk.Frame(self.frame_main)  # ボタンを表示するフレーム
        self.frame_button.pack(side="top")

        self.frame_checkbox = tk.Frame(self.frame_main)  # チェックボックスを表示するフレーム
        self.frame_checkbox.pack(side="top")

        self.frame_checkbox_lower = tk.Frame(self.frame_main)  # チェックボックスの下を表示するフレーム
        self.frame_checkbox_lower.pack(side="top")

        self.entry = tk.Entry(self.frame_entry)
        self.entry.pack(side="top")

        self.button1 = tk.Button(self.frame_button)
        self.button1["command"] = self.answer
        self.button1["text"] = "解答する"
        self.button1.pack(side="left", padx=20)

        self.button2 = tk.Button(self.frame_button)
        self.button2["command"] = self.giveup
        self.button2["text"] = "あきらめる"
        self.button2.pack(side="left", padx=20)

        self.button3 = tk.Button(self.frame_button)
        self.button3["command"] = lambda: [self.label1.destroy(), self.label2.destroy(), self.next()]
        self.button3["text"] = "次の問題へ"
        self.button3.pack(side="left", padx=20)

        self.boolean_var1 = tk.BooleanVar()
        self.boolean_var1.set(False)
        self.checkbutton1 = tk.Checkbutton(self.frame_checkbox)
        self.checkbutton1["variable"] = self.boolean_var1
        self.checkbutton1["text"] = "正答率50%以下の問題を表示"
        self.checkbutton1.pack(side="top")

        self.boolean_var2 = tk.BooleanVar()
        self.boolean_var2.set(False)
        self.checkbutton2 = tk.Checkbutton(self.frame_checkbox)
        self.checkbutton2["variable"] = self.boolean_var2
        self.checkbutton2["text"] = "解答回数5回以下の問題を表示"
        self.checkbutton2.pack(side="top")

        self.frame_sub = tk.Frame(self)
        self.frame_sub.grid(row=0, column=0, sticky="nsew")

        self.frame_text_sub = tk.Frame(self.frame_sub)  # 説明文を表示するフレーム
        self.frame_text_sub.pack(side="top")

        self.frame_button_sub = tk.Frame(self.frame_sub)  # ボタンを表示するフレーム
        self.frame_button_sub.pack(side="top")

        self.frame_checkbox_sub = tk.Frame(self.frame_sub)
        self.frame_checkbox_sub.pack(side="top")

        self.text_sub = tk.Label(self.frame_text_sub)
        self.text_sub["text"] = "最後の問題まで到達しました"
        self.text_sub.pack()

        self.button4 = tk.Button(self.frame_button_sub)
        self.button4["command"] = lambda: [self.label1.destroy(), self.label2.destroy(), self.reset()]
        self.button4["text"] = "初めから"
        self.button4.pack(side="left", padx=20)

        self.button5 = tk.Button(self.frame_button_sub)
        self.button5["command"] = self.quit
        self.button5["text"] = "終わる"
        self.button5.pack(side="left", padx=20)

        self.checkbutton3 = tk.Checkbutton(self.frame_checkbox_sub)
        self.checkbutton3["variable"] = self.boolean_var1
        self.checkbutton3["text"] = "正答率50%以下の問題を表示"
        self.checkbutton3.pack(side="top")

        self.checkbutton4 = tk.Checkbutton(self.frame_checkbox_sub)
        self.checkbutton4["variable"] = self.boolean_var2
        self.checkbutton4["text"] = "解答回数5回以下の問題を表示"
        self.checkbutton4.pack(side="top")

        self.frame_main.tkraise()

    def answer(self):
        text = self.entry.get()
        if App.sheet[self.row][WORD] == text:  # 正解したとき
            print("Correct!")
            if hasattr(self, "label2"):  # オブジェクトがlabel2属性を持っているとき、
                self.label2.destroy()  # label2属性を削除する
            self.label2 = tk.Label(self.frame_checkbox_lower, text="正解!", fg="red")
            self.label2.pack(side="top")
            self.button1["state"] = tk.DISABLED
            self.button2["state"] = tk.DISABLED
            self.button3["state"] = tk.NORMAL
            App.sheet[self.row][CORRECT] += 1
        else:  # 間違えたとき
            print("Incorrect!")
            if hasattr(self, "label2"):
                self.label2.destroy()
            self.label2 = tk.Label(self.frame_checkbox_lower, text="不正解!")
            self.label2.pack(side="top")
        App.sheet[self.row][TRY] += 1

    def giveup(self):
        print("Unanswered.")
        if hasattr(self, "label2"):
            self.label2.destroy()
        self.entry.delete(0, tk.END)  # 0文字目から最後の文字まで（つまり文字列のすべて）を削除する
        self.entry.insert(tk.END, App.sheet[self.row][WORD])
        self.label2 = tk.Label(self.frame_checkbox_lower, text="未解答")
        self.label2.pack(side="top")
        self.button1["state"] = tk.DISABLED
        self.button2["state"] = tk.DISABLED
        self.button3["state"] = tk.NORMAL
        App.sheet[self.row][TRY] += 1

    def next(self):
        wb = load_workbook(filename=FILENAME + ".xlsx")  # ファイルへの書き込みを行う
        sheet = wb[SHEETNAME]
        for i, j in itertools.product(range(0, len(App.sheet)), range(TRY, CORRECT + 1)):
            sheet.cell(row=i + 2, column=j + 1).value = App.sheet[i][j]
        try:
            wb.save(FILENAME + ".xlsx")
        except PermissionError:  # ファイルに上書きできないとき、エラーを受け取る
            App.permission_error = True  # エラー受け取ったとき、フラグを立てる

        while True:  # 表示する問題を選択
            if len(App.row_choice) == 0:  # 選択肢がないとき、
                print("List length is zero.")
                self.frame_sub.tkraise()  # 画面を遷移する
                return  # メソッドを終了する
                # break  # ループを抜ける

            self.row = random.choice(App.row_choice)  # ランダムに選択する
            App.row_choice.remove(self.row)  # 選択肢を削除する

            # 「正答率50%以下の問題を表示」に✓かつ、解答回数がゼロ回のとき、
            if (self.boolean_var1.get() and App.sheet[self.row][TRY] == 0
                # もしくは、「正答率50%以下の問題を表示」に✓かつ、正答率が50%以下のとき、
                or self.boolean_var1.get() and App.sheet[self.row][CORRECT] / App.sheet[self.row][TRY] <= 0.5
                # あるいは、「解答回数5回以下の問題を表示」に✓かつ、解答回数が5回以下のとき、
                or self.boolean_var2.get() and App.sheet[self.row][TRY] <= 5
                # あるいは、いずれの選択肢にも✓をしていないとき、
                or self.boolean_var1.get() is False and self.boolean_var2.get() is False):
                description = App.sheet[self.row][DESCRIPTION]  # 問題文をセットする
                break  # ループを抜ける

        print("Row number is " + str(self.row) + ".", 'Question is "' + description + '".')
        self.button1["state"] = tk.NORMAL if not App.permission_error else tk.DISABLED
        self.button2["state"] = tk.NORMAL if not App.permission_error else tk.DISABLED
        self.button3["state"] = tk.DISABLED
        self.entry.delete(0, tk.END)
        self.label1 = tk.Label(self.frame_text, text=description, wraplength=280) if not App.permission_error \
            else tk.Label(self.frame_text, text="開いているファイルを閉じて、アプリを再起動してください", fg="red", wraplength=280)
        self.label1.pack(side="top")

    def reset(self):
        App.row_choice = copy.deepcopy(App.row_list)
        self.frame_main.tkraise()
        self.next()

    def quit(self):
        global root
        root.quit()


def main():
    global root
    root = tk.Tk()
    app = App(master=root)
    app.mainloop()


if __name__ == "__main__":
    main()
